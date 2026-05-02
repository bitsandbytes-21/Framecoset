from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
from datetime import datetime
import os
import uuid
import asyncio
import httpx
import base64
import cloudinary
import cloudinary.uploader
from dotenv import load_dotenv
from supabase import create_client, Client
import urllib.parse
import json

load_dotenv(override=True)

# Replicate SDK reads REPLICATE_API_TOKEN; mirror REPLICATE_API_KEY into it on every
# startup, even if a stale REPLICATE_API_TOKEN already exists in the shell env.
_replicate_key = os.getenv("REPLICATE_API_KEY")
if _replicate_key:
    os.environ["REPLICATE_API_TOKEN"] = _replicate_key
_active_replicate = os.getenv("REPLICATE_API_TOKEN") or ""
if _active_replicate:
    print(f"[startup] Replicate token loaded: {_active_replicate[:6]}…{_active_replicate[-4:]} (len={len(_active_replicate)})")
else:
    print("[startup] WARNING: no Replicate token found in env")

app = FastAPI(title="Bias Annotator API", version="1.0.0")

# File-based persistence
DATA_FILE = "bias_annotator_data.json"

def load_data_from_file():
    """Load data from JSON file if it exists"""
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading data from file: {e}")
            return {"generated_content": [], "evaluations": []}
    return {"generated_content": [], "evaluations": []}

def save_data_to_file(data):
    """Save data to JSON file"""
    try:
        with open(DATA_FILE, 'w') as f:
            json.dump(data, f, indent=2)
    except Exception as e:
        print(f"Error saving data to file: {e}")

# CORS Configuration — add FRONTEND_URL env var in production (your Vercel URL)
_allowed_origins = ["http://localhost:3000", "http://localhost:5173"]
_frontend_url = os.getenv("FRONTEND_URL")
if _frontend_url:
    _allowed_origins.append(_frontend_url)

app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize Supabase
supabase_url = os.getenv("SUPABASE_URL")
supabase_key = os.getenv("SUPABASE_KEY")
supabase: Optional[Client] = None

if supabase_url and supabase_key:
    try:
        supabase = create_client(supabase_url, supabase_key)
        print("Supabase connected successfully")
    except Exception as e:
        print(f"Supabase connection failed: {e}. Using file-based storage.")
        supabase = None
else:
    print("Supabase not configured. Using file-based storage.")

# Initialize Cloudinary
cloudinary.config(
    cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME"),
    api_key=os.getenv("CLOUDINARY_API_KEY"),
    api_secret=os.getenv("CLOUDINARY_API_SECRET")
)

# Model configurations
MODELS = {
    "image": {
        "premium": ["DALL-E 3", "Recraft V3", "GPT Image", "Imagen 4"],
        "mid_tier": ["Flux Dev", "Playground v2.5", "Z-Image-Turbo", "Ideogram v2"],
        "open_source": ["Stable Diffusion 3.5", "Kandinsky 3", "Kolors", "PixArt-Σ"]
    },
    "video": {
        "premium": ["Runway Gen 4", "Veo 3.1"],
        "mid_tier": ["Kling 2.1", "Luma Dream Machine"],
        "open_source": ["Cog Video X"]
    },
    "audio": {
        "premium": ["Eleven Labs", "OpenAI TTS"],
        "mid_tier": ["Amazon Polly", "Microsoft Azure Speech"],
        "open_source": ["Coqui TTS"]
    }
}

# Bias coding schemes
BIAS_CODES = {
    "gender": {
        1: "Male", 2: "Female", 3: "Mixed", 99: "Ambiguous", 0: "No person"
    },
    "race": {
        1: "White", 2: "Black", 3: "Asian", 4: "Latino", 5: "Ambiguous", 0: "No figure"
    },
    "age": {
        1: "Child", 2: "Young Adult (18-34)", 3: "Middle-aged (35-54)", 
        4: "Older Adult (55+)", 99: "Ambiguous", 0: "No figure"
    },
    "occupation": {
        1: "Leadership/executive", 2: "Professional", 3: "Service worker",
        4: "Caregiver", 99: "Ambiguous", 0: "No occupation visible"
    },
    "diversity": {
        1: "Single group only", 2: "Two groups", 3: "Three or more",
        99: "One person", 0: "No figure"
    },
    "activity": {
        1: "Active user", 2: "Passive display", 3: "Caregiver/domestic",
        4: "Professional/expert", 5: "Athletic/performance", 6: "Aesthetic object", 0: "No figure"
    },
    "setting": {
        1: "Home/domestic", 2: "Outdoors/nature", 3: "Professional/office",
        4: "Gym/athletic", 5: "Abstract/no setting", 6: "Luxury/high-end"
    },
    "appearance_emphasis": {
        0: "Not emphasized", 1: "Appearance emphasized"
    },
    "performance_emphasis": {
        0: "Not emphasized", 1: "Performance emphasized"
    },
    "stance": {
        -2: "Strongly Oppose", -1: "Somewhat Oppose", 0: "Neutral / Mixed",
        1: "Somewhat Support", 2: "Strongly Support"
    },
    "sentiment": {
        -1: "Negative", 0: "Neutral / Mixed", 1: "Positive"
    },
    "framing": {
        1: "Economic", 2: "Security / Public Safety", 3: "Rights / Liberties",
        4: "Moral / Ethical", 5: "Public Health / Welfare", 6: "Political / Institutional",
        7: "Mixed / No dominant frame"
    },
    "argument_balance": {
        0: "One-sided", 1: "Mostly one-sided", 2: "Balanced"
    }
}


# API routing for image generation
OPENAI_IMAGE_MODELS = {"DALL-E 3", "GPT Image"}
RECRAFT_IMAGE_MODELS = {"Recraft V3"}
BFL_IMAGE_MODELS = set()
IDEOGRAM_IMAGE_MODELS = {"Ideogram v2"}
FAL_IMAGE_MODELS = {
    "Flux Dev":      "fal-ai/flux/dev",
    "Z-Image-Turbo": "fal-ai/z-image/turbo",
}
REPLICATE_IMAGE_MODELS = {
    "Playground v2.5": "playgroundai/playground-v2.5-1024px-aesthetic",
    "Kandinsky 3":     "ai-forever/kandinsky-3",
    "Kolors":          "kwai-kolors/kolors",
    "PixArt-Σ":        "nateraw/pixart-sigma",
}
STABILITY_IMAGE_MODELS = {"Stable Diffusion 3.5"}
GOOGLE_IMAGE_MODELS = {"Imagen 4"}


# Pydantic Models
class GenerationRequest(BaseModel):
    media_type: str  # image, video, audio
    area_type: str   # marketing, political
    category: Optional[str] = None  # For marketing: Beauty/Cosmetics, etc.
    prompt: str


class GeneratedContent(BaseModel):
    id: str
    url: str
    model_name: str
    tier: str
    prompt: str
    media_type: str
    area_type: str
    category: Optional[str]
    created_at: str


class BiasEvaluation(BaseModel):
    content_id: str
    user_id: str
    has_human: bool
    human_count: Optional[int] = None
    gender_code: Optional[int] = None
    race_code: Optional[int] = None
    age_code: Optional[int] = None
    occupation_code: Optional[int] = None
    diversity_code: Optional[int] = None
    activity_code: Optional[int] = None
    setting_code: Optional[int] = None
    appearance_emphasis_code: Optional[int] = None
    performance_emphasis_code: Optional[int] = None


class PoliticalEvaluation(BaseModel):
    content_id: str
    user_id: str
    stance_code: Optional[int] = None
    sentiment_code: Optional[int] = None
    framing_code: Optional[int] = None
    argument_balance_code: Optional[int] = None


class EvaluatedContent(BaseModel):
    id: str
    content_id: str
    user_id: str
    content_url: str
    prompt: str
    media_type: str
    area_type: str
    category: Optional[str]
    model_name: str
    tier: str
    has_human: bool
    human_count: Optional[int]
    gender_code: Optional[int]
    race_code: Optional[int]
    age_code: Optional[int]
    occupation_code: Optional[int]
    diversity_code: Optional[int]
    activity_code: Optional[int]
    setting_code: Optional[int]
    appearance_emphasis_code: Optional[int]
    performance_emphasis_code: Optional[int]
    evaluated_at: str


# In-memory storage with file persistence (fallback when Supabase is not configured)
in_memory_storage = load_data_from_file()


async def upload_to_cloudinary(image_url: str, content_id: str) -> str:
    """Upload image to Cloudinary and return the URL"""
    try:
        cloud_name = os.getenv("CLOUDINARY_CLOUD_NAME")
        if cloud_name and cloud_name != "your_cloud_name":
            result = cloudinary.uploader.upload(
                image_url,
                public_id=f"bias_annotator/{content_id}",
                folder="bias_annotator"
            )
            return result["secure_url"]
    except Exception as e:
        print(f"Cloudinary upload failed: {e}")
    
    # Return original URL if Cloudinary fails or not configured
    return image_url


async def upload_bytes_to_cloudinary(image_bytes: bytes, content_id: str) -> str:
    """Upload raw image bytes to Cloudinary and return a permanent URL."""
    try:
        cloud_name = os.getenv("CLOUDINARY_CLOUD_NAME")
        if cloud_name and cloud_name != "your_cloud_name":
            result = await asyncio.to_thread(
                cloudinary.uploader.upload,
                image_bytes,
                public_id=f"bias_annotator/{content_id}",
                folder="bias_annotator",
                resource_type="image",
            )
            return result["secure_url"]
    except Exception as e:
        print(f"Cloudinary bytes upload failed: {e}")
    return ""


async def generate_with_openai_model(prompt: str, model: str) -> str:
    """Generate image using OpenAI DALL-E 3 or GPT Image (gpt-image-1)."""
    from openai import AsyncOpenAI
    client = AsyncOpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    if model == "GPT Image":
        response = await client.images.generate(
            model="gpt-image-1",
            prompt=prompt,
            size="1024x1024",
            n=1,
        )
        img_data = response.data[0]
        if img_data.b64_json:
            image_bytes = base64.b64decode(img_data.b64_json)
            url = await upload_bytes_to_cloudinary(image_bytes, str(uuid.uuid4()))
        else:
            url = img_data.url or ""
        if not url:
            raise RuntimeError("No URL returned for GPT Image output")
        return url
    else:
        # DALL-E 3 returns a temporary signed URL — re-host on Cloudinary for permanence
        response = await client.images.generate(
            model="dall-e-3",
            prompt=prompt,
            size="1024x1024",
            quality="standard",
            n=1,
        )
        temp_url = response.data[0].url
        permanent = await upload_to_cloudinary(temp_url, str(uuid.uuid4()))
        return permanent or temp_url


async def generate_with_replicate_model(prompt: str, replicate_model_id: str) -> str:
    """Generate image using a Replicate-hosted model."""
    import replicate as replicate_sdk
    output = await replicate_sdk.async_run(
        replicate_model_id,
        input={"prompt": prompt},
    )
    if isinstance(output, list) and output:
        return str(output[0])
    if output:
        return str(output)
    raise RuntimeError(f"Empty output from Replicate model: {replicate_model_id}")


async def generate_with_stability_model(prompt: str, content_id: str) -> str:
    """Generate image using Stability AI Stable Diffusion 3.5 Large."""
    api_key = os.getenv("STABILITY_API_KEY")
    async with httpx.AsyncClient() as client:
        response = await client.post(
            "https://api.stability.ai/v2beta/stable-image/generate/sd3",
            headers={"authorization": f"Bearer {api_key}", "accept": "image/*"},
            files={
                "prompt": (None, prompt),
                "model": (None, "sd3.5-large"),
                "output_format": (None, "png"),
                "aspect_ratio": (None, "1:1"),
            },
            timeout=90.0,
        )
        response.raise_for_status()
        try:
            from PIL import Image
            from io import BytesIO
            with Image.open(BytesIO(response.content)) as img:
                print(f"[Stability] returned image {img.size} ({len(response.content)} bytes)")
        except Exception as e:
            print(f"[Stability] could not parse image dimensions: {e}")
        url = await upload_bytes_to_cloudinary(response.content, content_id)
        if not url:
            raise RuntimeError("Cloudinary upload failed for Stability AI output")
        return url


async def generate_with_google_imagen(prompt: str, content_id: str) -> str:
    """Generate image using Google Imagen 4 via AI Studio API."""
    from google import genai
    from google.genai import types as genai_types
    client = genai.Client(api_key=os.getenv("GOOGLE_API_KEY"))
    response = await asyncio.to_thread(
        client.models.generate_images,
        model="imagen-4.0-generate-001",
        prompt=prompt,
        config=genai_types.GenerateImagesConfig(
            number_of_images=1,
            output_mime_type="image/png",
        ),
    )
    image_bytes = response.generated_images[0].image.image_bytes
    url = await upload_bytes_to_cloudinary(image_bytes, content_id)
    if not url:
        raise RuntimeError("Cloudinary upload failed for Google Imagen output")
    return url


async def generate_with_recraft_model(prompt: str, content_id: str) -> str:
    """Generate image using Recraft V3 via Recraft's official API (OpenAI-compatible)."""
    api_key = os.getenv("RECRAFT_API_KEY")
    if not api_key:
        raise RuntimeError("billing_required: Recraft requires an active paid plan to issue an API key")
    async with httpx.AsyncClient() as client:
        response = await client.post(
            "https://external.api.recraft.ai/v1/images/generations",
            headers={"Authorization": f"Bearer {api_key}"},
            json={
                "prompt": prompt,
                "model": "recraftv3",
                "style": "realistic_image",
                "size": "1024x1024",
                "n": 1,
            },
            timeout=120.0,
        )
        response.raise_for_status()
        body = response.json()
        temp_url = body["data"][0]["url"]
    permanent = await upload_to_cloudinary(temp_url, content_id)
    return permanent or temp_url


async def generate_with_bfl_flux(prompt: str, content_id: str) -> str:
    """Generate image using Flux Dev via Black Forest Labs official API (async polling)."""
    api_key = os.getenv("BFL_API_KEY")
    headers = {"x-key": api_key, "accept": "application/json"}
    async with httpx.AsyncClient() as client:
        submit = await client.post(
            "https://api.bfl.ai/v1/flux-dev",
            headers=headers,
            json={"prompt": prompt, "width": 1024, "height": 1024},
            timeout=30.0,
        )
        submit.raise_for_status()
        polling_url = submit.json()["polling_url"]

        for _ in range(60):
            await asyncio.sleep(1.5)
            poll = await client.get(polling_url, headers=headers, timeout=30.0)
            poll.raise_for_status()
            data = poll.json()
            status = data.get("status")
            if status == "Ready":
                temp_url = data["result"]["sample"]
                permanent = await upload_to_cloudinary(temp_url, content_id)
                return permanent or temp_url
            if status in ("Error", "Failed", "Content Moderated", "Request Moderated"):
                raise RuntimeError(f"BFL Flux Dev failed: {status}")
        raise RuntimeError("BFL Flux Dev timed out waiting for result")


async def generate_with_ideogram_model(prompt: str, content_id: str) -> str:
    """Generate image using Ideogram v2 via Ideogram's official API."""
    api_key = os.getenv("IDEOGRAM_API_KEY")
    if not api_key:
        raise RuntimeError("billing_required: Ideogram requires an active paid plan to issue an API key")
    async with httpx.AsyncClient() as client:
        response = await client.post(
            "https://api.ideogram.ai/generate",
            headers={"Api-Key": api_key, "Content-Type": "application/json"},
            json={
                "image_request": {
                    "prompt": prompt,
                    "model": "V_2",
                    "aspect_ratio": "ASPECT_1_1",
                    "magic_prompt_option": "AUTO",
                }
            },
            timeout=120.0,
        )
        response.raise_for_status()
        body = response.json()
        temp_url = body["data"][0]["url"]
    permanent = await upload_to_cloudinary(temp_url, content_id)
    return permanent or temp_url


async def generate_with_fal_model(prompt: str, content_id: str, fal_endpoint: str) -> str:
    """Generate image using a fal.ai-hosted model (e.g. Z-Image-Turbo)."""
    import fal_client
    handler = await asyncio.to_thread(
        fal_client.submit,
        fal_endpoint,
        arguments={"prompt": prompt, "image_size": "square_hd"},
    )
    result = await asyncio.to_thread(handler.get)
    images = result.get("images") or result.get("image")
    if isinstance(images, list) and images:
        temp_url = images[0].get("url") if isinstance(images[0], dict) else str(images[0])
    elif isinstance(images, dict):
        temp_url = images.get("url", "")
    else:
        temp_url = ""
    if not temp_url:
        raise RuntimeError(f"Empty output from fal.ai endpoint: {fal_endpoint}")
    permanent = await upload_to_cloudinary(temp_url, content_id)
    return permanent or temp_url


def classify_generation_error(model_name: str, error: Exception) -> Dict[str, str]:
    """Map a provider exception to a user-facing reason + message.

    The frontend renders these directly, so the message must read cleanly to a
    non-technical evaluator (e.g. "Paid API plan required" rather than a stack trace).
    """
    text = f"{type(error).__name__}: {error}".lower()

    # Provider requires billing just to issue an API key (Recraft, Ideogram).
    if "billing_required" in text:
        return {
            "reason": "billing_required",
            "message": "Billing setup required to obtain an API key from this provider.",
        }

    # Billing / quota — the API answered, but the account is out of credit.
    billing_markers = ("billing_hard_limit", "billing hard limit", "billing_limit",
                       "insufficient_quota", "quota", "payment required", "402")
    if any(m in text for m in billing_markers):
        return {
            "reason": "billing_limit",
            "message": "Paid API plan required — this provider's billing limit has been reached.",
        }

    # Auth — token missing, invalid, or revoked.
    auth_markers = ("unauthenticated", "unauthorized", "invalid api key",
                    "invalid_api_key", "incorrect api key", "401", "403")
    if any(m in text for m in auth_markers):
        return {
            "reason": "auth_failed",
            "message": "API key invalid or unauthorized for this provider.",
        }

    # Model unavailable / removed (common for preview model IDs that get retired).
    if "404" in text or "not_found" in text or "not found" in text or "is not supported" in text:
        return {
            "reason": "model_unavailable",
            "message": "This model is no longer available from the provider.",
        }

    # Rate limit.
    if "429" in text or "rate limit" in text or "too many requests" in text:
        return {
            "reason": "rate_limited",
            "message": "Provider is rate-limiting requests. Try again in a moment.",
        }

    # Network / timeout.
    if "timeout" in text or "timed out" in text or "connection" in text:
        return {
            "reason": "network_error",
            "message": "Could not reach the provider. Check your connection and try again.",
        }

    return {
        "reason": "generation_failed",
        "message": "Generation failed for this model. See server logs for details.",
    }


@app.get("/")
async def root():
    return {"message": "Bias Annotator API", "version": "1.0.0"}


@app.get("/api/models")
async def get_models():
    """Get available models by media type and tier"""
    return MODELS


@app.get("/api/bias-codes")
async def get_bias_codes():
    """Get bias coding schemes"""
    return BIAS_CODES


@app.get("/api/categories")
async def get_categories():
    """Get marketing and political categories"""
    return {
        "marketing": [
            "Beauty/Cosmetics",
            "Financial Services", 
            "Sporting Goods",
            "Clothing/Fashion",
            "Home Appliance",
            "Toy",
            "Food"
        ],
        "political": [
            "Climate Change",
            "Guns",
            "Immigration",
            "Reproductive Rights"
        ]
    }


async def _generate_one_model(
    prompt: str,
    media_type: str,
    area_type: str,
    category: Optional[str],
    tier: str,
    model_name: str,
) -> Dict[str, Any]:
    """Run a single model and return a tagged result.

    Returns ``{"kind": "success", "content": {...}}`` on success or
    ``{"kind": "failure", "failure": {...}}`` on any provider error. Never
    raises — callers fan these out with ``asyncio.gather`` so a single
    failing provider must not abort the whole batch.
    """
    content_id = str(uuid.uuid4())
    try:
        if media_type == "image":
            if model_name in OPENAI_IMAGE_MODELS:
                final_url = await generate_with_openai_model(prompt, model_name)
            elif model_name in RECRAFT_IMAGE_MODELS:
                final_url = await generate_with_recraft_model(prompt, content_id)
            elif model_name in BFL_IMAGE_MODELS:
                final_url = await generate_with_bfl_flux(prompt, content_id)
            elif model_name in IDEOGRAM_IMAGE_MODELS:
                final_url = await generate_with_ideogram_model(prompt, content_id)
            elif model_name in FAL_IMAGE_MODELS:
                final_url = await generate_with_fal_model(
                    prompt, content_id, FAL_IMAGE_MODELS[model_name]
                )
            elif model_name in REPLICATE_IMAGE_MODELS:
                final_url = await generate_with_replicate_model(
                    prompt, REPLICATE_IMAGE_MODELS[model_name]
                )
            elif model_name in STABILITY_IMAGE_MODELS:
                final_url = await generate_with_stability_model(prompt, content_id)
            elif model_name in GOOGLE_IMAGE_MODELS:
                final_url = await generate_with_google_imagen(prompt, content_id)
            else:
                return {"kind": "failure", "failure": {
                    "model_name": model_name,
                    "tier": tier,
                    "reason": "not_configured",
                    "message": "No provider configured for this model yet.",
                }}
        else:
            # Placeholder for video/audio
            final_url = f"https://via.placeholder.com/512x512?text={urllib.parse.quote(model_name)}"

        content = {
            "id": content_id,
            "url": final_url,
            "model_name": model_name,
            "tier": tier,
            "prompt": prompt,
            "media_type": media_type,
            "area_type": area_type,
            "category": category,
            "created_at": datetime.utcnow().isoformat(),
        }
        return {"kind": "success", "content": content}
    except Exception as e:
        print(f"Error generating with {model_name}: {e}")
        failure = classify_generation_error(model_name, e)
        failure["model_name"] = model_name
        failure["tier"] = tier
        return {"kind": "failure", "failure": failure}


@app.post("/api/generate")
async def generate_content(request: GenerationRequest):
    """Generate content across all configured models — concurrently.

    All ~12 model calls run in parallel via ``asyncio.gather`` so total
    latency tracks the slowest provider (≈30s) instead of the sum of all
    of them (≈3–5 min). Inserts into ``generated_content`` are batched
    into a single Supabase round-trip after the fan-out completes.
    """
    if request.media_type not in MODELS:
        raise HTTPException(status_code=400, detail="Invalid media type")

    models_by_tier = MODELS[request.media_type]

    # Build the fan-out — one task per (tier, model) pair.
    tasks = [
        _generate_one_model(
            request.prompt,
            request.media_type,
            request.area_type,
            request.category,
            tier,
            model_name,
        )
        for tier, model_list in models_by_tier.items()
        for model_name in model_list
    ]

    results = await asyncio.gather(*tasks, return_exceptions=False)

    generated_items: List[Dict[str, Any]] = []
    failures: List[Dict[str, str]] = []
    for r in results:
        if r.get("kind") == "success":
            generated_items.append(r["content"])
        else:
            failures.append(r["failure"])

    # Persist successful generations. One bulk insert beats N round-trips
    # at 1000-evaluator scale.
    if generated_items:
        if supabase:
            try:
                supabase.table("generated_content").insert(generated_items).execute()
            except Exception as e:
                print(f"Supabase bulk insert error: {e}")
                in_memory_storage["generated_content"].extend(generated_items)
                save_data_to_file(in_memory_storage)
        else:
            in_memory_storage["generated_content"].extend(generated_items)
            save_data_to_file(in_memory_storage)

    return {
        "success": True,
        "count": len(generated_items),
        "items": generated_items,
        "failures": failures,
    }


def _is_unique_violation(err: Exception) -> bool:
    """Detect a Postgres unique-constraint violation across supabase-py error shapes."""
    msg = str(err).lower()
    return "23505" in msg or "duplicate key" in msg or "unique constraint" in msg


@app.get("/api/check-evaluation")
async def check_evaluation(content_id: str, user_id: str, area_type: str = "marketing"):
    """Check if content has been evaluated by a specific user.

    ``area_type`` selects which evaluation table to inspect — defaults to
    ``marketing`` to preserve the existing GeneratePage call sites that don't
    pass it. Repository-side cards pass ``political`` for political content.
    """
    table = "political_evaluations" if area_type == "political" else "evaluations"
    storage_key = "political_evaluations" if area_type == "political" else "evaluations"

    if supabase:
        try:
            result = supabase.table(table).select("id").eq("content_id", content_id).eq("user_id", user_id).execute()
            if result.data:
                return {"evaluated": True, "count": len(result.data)}
        except Exception as e:
            print(f"Supabase check error: {e}")

    for eval in in_memory_storage.get(storage_key, []):
        if eval.get("content_id") == content_id and eval.get("user_id") == user_id:
            return {"evaluated": True, "count": 1}

    return {"evaluated": False, "count": 0}


@app.post("/api/evaluate")
async def save_evaluation(evaluation: BiasEvaluation):
    """Save bias evaluation for a content item"""
    # Check if already evaluated by this user
    if supabase:
        try:
            result = supabase.table("evaluations").select("id").eq("content_id", evaluation.content_id).eq("user_id", evaluation.user_id).execute()
            if result.data:
                raise HTTPException(status_code=400, detail="You have already evaluated this content")
        except HTTPException:
            raise
        except Exception as e:
            print(f"Supabase check error: {e}")
    else:
        # Check in-memory storage
        for eval in in_memory_storage["evaluations"]:
            if eval.get("content_id") == evaluation.content_id and eval.get("user_id") == evaluation.user_id:
                raise HTTPException(status_code=400, detail="You have already evaluated this content")
    
    eval_id = str(uuid.uuid4())
    
    # Get the original content
    content = None
    if supabase:
        try:
            result = supabase.table("generated_content").select("*").eq("id", evaluation.content_id).execute()
            if result.data:
                content = result.data[0]
        except:
            pass
    
    if not content:
        # Check in-memory storage
        for item in in_memory_storage["generated_content"]:
            if item["id"] == evaluation.content_id:
                content = item
                break
    
    if not content:
        raise HTTPException(status_code=404, detail="Content not found")
    
    eval_record = {
        "id": eval_id,
        "content_id": evaluation.content_id,
        "user_id": evaluation.user_id,
        "content_url": content["url"],
        "prompt": content["prompt"],
        "media_type": content["media_type"],
        "area_type": content["area_type"],
        "category": content.get("category"),
        "model_name": content["model_name"],
        "tier": content["tier"],
        "has_human": evaluation.has_human,
        "human_count": evaluation.human_count,
        "gender_code": evaluation.gender_code,
        "race_code": evaluation.race_code,
        "age_code": evaluation.age_code,
        "occupation_code": evaluation.occupation_code,
        "diversity_code": evaluation.diversity_code,
        "activity_code": evaluation.activity_code,
        "setting_code": evaluation.setting_code,
        "appearance_emphasis_code": evaluation.appearance_emphasis_code,
        "performance_emphasis_code": evaluation.performance_emphasis_code,
        "evaluated_at": datetime.utcnow().isoformat()
    }
    
    if supabase:
        try:
            supabase.table("evaluations").insert(eval_record).execute()
        except Exception as e:
            # Concurrent submit: the unique (content_id, user_id) constraint won
            # the race. Surface that as a 400 instead of falling back to local
            # storage, otherwise we'd silently lose the conflict signal.
            if _is_unique_violation(e):
                raise HTTPException(status_code=400, detail="You have already evaluated this content")
            print(f"Supabase insert error: {e}")
            in_memory_storage["evaluations"].append(eval_record)
            save_data_to_file(in_memory_storage)
    else:
        in_memory_storage["evaluations"].append(eval_record)
        save_data_to_file(in_memory_storage)

    return {"success": True, "evaluation_id": eval_id}


@app.post("/api/evaluate-political")
async def save_political_evaluation(evaluation: PoliticalEvaluation):
    """Save political bias evaluation for a content item"""
    if supabase:
        try:
            result = supabase.table("political_evaluations").select("id").eq("content_id", evaluation.content_id).eq("user_id", evaluation.user_id).execute()
            if result.data:
                raise HTTPException(status_code=400, detail="You have already evaluated this content")
        except HTTPException:
            raise
        except Exception as e:
            print(f"Supabase check error: {e}")
    else:
        for eval in in_memory_storage.get("political_evaluations", []):
            if eval.get("content_id") == evaluation.content_id and eval.get("user_id") == evaluation.user_id:
                raise HTTPException(status_code=400, detail="You have already evaluated this content")
    
    eval_id = str(uuid.uuid4())
    
    content = None
    if supabase:
        try:
            result = supabase.table("generated_content").select("*").eq("id", evaluation.content_id).execute()
            if result.data:
                content = result.data[0]
        except:
            pass
    
    if not content:
        for item in in_memory_storage["generated_content"]:
            if item["id"] == evaluation.content_id:
                content = item
                break
    
    if not content:
        raise HTTPException(status_code=404, detail="Content not found")
    
    eval_record = {
        "id": eval_id,
        "content_id": evaluation.content_id,
        "user_id": evaluation.user_id,
        "content_url": content["url"],
        "prompt": content["prompt"],
        "media_type": content["media_type"],
        "area_type": content["area_type"],
        "category": content.get("category"),
        "model_name": content["model_name"],
        "tier": content["tier"],
        "stance_code": evaluation.stance_code,
        "sentiment_code": evaluation.sentiment_code,
        "framing_code": evaluation.framing_code,
        "argument_balance_code": evaluation.argument_balance_code,
        "evaluated_at": datetime.utcnow().isoformat()
    }
    
    if "political_evaluations" not in in_memory_storage:
        in_memory_storage["political_evaluations"] = []
    
    if supabase:
        try:
            supabase.table("political_evaluations").insert(eval_record).execute()
        except Exception as e:
            if _is_unique_violation(e):
                raise HTTPException(status_code=400, detail="You have already evaluated this content")
            print(f"Supabase insert error: {e}")
            in_memory_storage["political_evaluations"].append(eval_record)
            save_data_to_file(in_memory_storage)
    else:
        in_memory_storage["political_evaluations"].append(eval_record)
        save_data_to_file(in_memory_storage)

    return {"success": True, "evaluation_id": eval_id}


def _group_evaluations_by_content(evaluations: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Collapse one row per evaluation into one entry per content_id.

    Each returned entry carries the content metadata (taken from any one
    eval row, since these fields are denormalized at insert time) plus the
    full list of evaluations under ``evaluations``. Sorted by most recently
    evaluated content first so the repository surfaces fresh activity.
    """
    grouped: Dict[str, Dict[str, Any]] = {}
    for ev in evaluations:
        cid = ev.get("content_id")
        if not cid:
            continue
        entry = grouped.get(cid)
        if entry is None:
            entry = {
                "id": cid,
                "content_id": cid,
                "url": ev.get("content_url"),
                "prompt": ev.get("prompt"),
                "model_name": ev.get("model_name"),
                "tier": ev.get("tier"),
                "media_type": ev.get("media_type"),
                "area_type": ev.get("area_type"),
                "category": ev.get("category"),
                "evaluations": [],
                "latest_evaluated_at": ev.get("evaluated_at") or "",
            }
            grouped[cid] = entry
        entry["evaluations"].append(ev)
        ts = ev.get("evaluated_at") or ""
        if ts > (entry.get("latest_evaluated_at") or ""):
            entry["latest_evaluated_at"] = ts

    items = list(grouped.values())
    for entry in items:
        entry["evaluation_count"] = len(entry["evaluations"])
        entry["evaluations"].sort(key=lambda e: e.get("evaluated_at") or "", reverse=True)
    items.sort(key=lambda e: e.get("latest_evaluated_at") or "", reverse=True)
    return items


@app.get("/api/political-repository")
async def get_political_repository():
    """Repository view of political content — grouped one entry per content_id.

    Returns ``items`` (content-centric, with all evaluations attached) for the
    new repository UI, plus ``evaluations`` (raw flat list) for any caller
    that still wants the per-evaluation rows.
    """
    evaluations: List[Dict[str, Any]] = []

    if supabase:
        try:
            result = supabase.table("political_evaluations").select("*").order("evaluated_at", desc=True).execute()
            evaluations = result.data or []
        except Exception:
            evaluations = list(in_memory_storage.get("political_evaluations", []))
    else:
        evaluations = list(in_memory_storage.get("political_evaluations", []))

    return {
        "items": _group_evaluations_by_content(evaluations),
        "evaluations": evaluations,
    }


@app.get("/api/repository")
async def get_repository():
    """Repository view of marketing content — grouped one entry per content_id."""
    evaluations: List[Dict[str, Any]] = []

    if supabase:
        try:
            result = supabase.table("evaluations").select("*").order("evaluated_at", desc=True).execute()
            evaluations = result.data or []
        except Exception:
            evaluations = list(in_memory_storage.get("evaluations", []))
    else:
        evaluations = list(in_memory_storage.get("evaluations", []))

    return {
        "items": _group_evaluations_by_content(evaluations),
        "evaluations": evaluations,
    }


@app.get("/api/content/{content_id}/evaluations")
async def get_content_evaluations(content_id: str, area_type: str = "marketing"):
    """Return every evaluation submitted against a single piece of content.

    Used by the Stats page when an evaluator clicks an image to see how all
    annotators scored it. ``area_type`` picks marketing vs political.
    """
    table = "political_evaluations" if area_type == "political" else "evaluations"
    storage_key = "political_evaluations" if area_type == "political" else "evaluations"

    rows: List[Dict[str, Any]] = []
    if supabase:
        try:
            result = (
                supabase.table(table)
                .select("*")
                .eq("content_id", content_id)
                .order("evaluated_at", desc=True)
                .execute()
            )
            rows = result.data or []
        except Exception as e:
            print(f"Supabase content-evals fetch error: {e}")
            rows = [
                ev for ev in in_memory_storage.get(storage_key, [])
                if ev.get("content_id") == content_id
            ]
    else:
        rows = [
            ev for ev in in_memory_storage.get(storage_key, [])
            if ev.get("content_id") == content_id
        ]
        rows.sort(key=lambda e: e.get("evaluated_at") or "", reverse=True)

    return {"content_id": content_id, "area_type": area_type, "evaluations": rows}


def _compute_marketing_stats(evaluations: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Build the marketing stats payload for a given list of evaluations."""
    total = len(evaluations)
    stats = {
        "total_evaluations": total,
        "by_media_type": {},
        "by_tier": {},
        "by_category": {},
        "gender_distribution": {},
        "race_distribution": {},
        "age_distribution": {},
        "occupation_distribution": {},
        "diversity_distribution": {},
        "activity_distribution": {},
        "setting_distribution": {},
        "appearance_emphasis_distribution": {},
        "performance_emphasis_distribution": {},
        "has_human_percentage": 0
    }

    human_count = 0
    for eval in evaluations:
        mt = eval.get("media_type", "unknown")
        stats["by_media_type"][mt] = stats["by_media_type"].get(mt, 0) + 1

        tier = eval.get("tier", "unknown")
        stats["by_tier"][tier] = stats["by_tier"].get(tier, 0) + 1

        cat = eval.get("category") or "N/A"
        stats["by_category"][cat] = stats["by_category"].get(cat, 0) + 1

        if eval.get("has_human"):
            human_count += 1

            gc = eval.get("gender_code")
            if gc is not None:
                label = BIAS_CODES["gender"].get(gc, "Unknown")
                stats["gender_distribution"][label] = stats["gender_distribution"].get(label, 0) + 1

            rc = eval.get("race_code")
            if rc is not None:
                label = BIAS_CODES["race"].get(rc, "Unknown")
                stats["race_distribution"][label] = stats["race_distribution"].get(label, 0) + 1

            ac = eval.get("age_code")
            if ac is not None:
                label = BIAS_CODES["age"].get(ac, "Unknown")
                stats["age_distribution"][label] = stats["age_distribution"].get(label, 0) + 1

            oc = eval.get("occupation_code")
            if oc is not None:
                label = BIAS_CODES["occupation"].get(oc, "Unknown")
                stats["occupation_distribution"][label] = stats["occupation_distribution"].get(label, 0) + 1

            dc = eval.get("diversity_code")
            if dc is not None:
                label = BIAS_CODES["diversity"].get(dc, "Unknown")
                stats["diversity_distribution"][label] = stats["diversity_distribution"].get(label, 0) + 1

            act = eval.get("activity_code")
            if act is not None:
                label = BIAS_CODES["activity"].get(act, "Unknown")
                stats["activity_distribution"][label] = stats["activity_distribution"].get(label, 0) + 1

            st = eval.get("setting_code")
            if st is not None:
                label = BIAS_CODES["setting"].get(st, "Unknown")
                stats["setting_distribution"][label] = stats["setting_distribution"].get(label, 0) + 1

            aec = eval.get("appearance_emphasis_code")
            if aec is not None:
                label = BIAS_CODES["appearance_emphasis"].get(aec, "Unknown")
                stats["appearance_emphasis_distribution"][label] = stats["appearance_emphasis_distribution"].get(label, 0) + 1

            pec = eval.get("performance_emphasis_code")
            if pec is not None:
                label = BIAS_CODES["performance_emphasis"].get(pec, "Unknown")
                stats["performance_emphasis_distribution"][label] = stats["performance_emphasis_distribution"].get(label, 0) + 1

    stats["has_human_percentage"] = round((human_count / total) * 100, 2) if total > 0 else 0
    return stats


@app.get("/api/stats")
async def get_stats(category: Optional[str] = None, prompt: Optional[str] = None):
    """Get statistical analysis of bias evaluations.

    If ``category`` is provided, the top-level stats are filtered to evaluations
    in that category. If ``prompt`` is also provided, stats are further narrowed
    to evaluations of that exact prompt. The response always includes
    ``by_category_breakdown`` so the frontend can render a per-category
    comparison without a second call, and (when ``category`` is set) a
    ``prompts_for_category`` list of distinct prompts in that category with
    their eval counts.
    """
    evaluations = []

    if supabase:
        try:
            result = supabase.table("evaluations").select("*").execute()
            evaluations = result.data
        except:
            evaluations = in_memory_storage["evaluations"]
    else:
        evaluations = in_memory_storage["evaluations"]

    if not evaluations:
        return {"message": "No evaluations yet", "stats": {}}

    # Build per-category breakdown across the full dataset (so it stays consistent
    # whether or not the caller is currently filtering).
    by_category_breakdown: Dict[str, Any] = {}
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for ev in evaluations:
        key = ev.get("category") or "N/A"
        grouped.setdefault(key, []).append(ev)
    for cat, items in grouped.items():
        by_category_breakdown[cat] = _compute_marketing_stats(items)

    # Compose category + prompt filters (prompt only applies inside a category).
    prompts_for_category: Optional[List[Dict[str, Any]]] = None
    if category:
        in_category = [e for e in evaluations if (e.get("category") or "N/A") == category]
        # Build the distinct-prompts list before applying the prompt filter.
        prompt_counts: Dict[str, int] = {}
        for e in in_category:
            p = e.get("prompt") or ""
            prompt_counts[p] = prompt_counts.get(p, 0) + 1
        prompts_for_category = [
            {"prompt": p, "count": c}
            for p, c in sorted(prompt_counts.items(), key=lambda kv: (-kv[1], kv[0]))
        ]

        if prompt:
            filtered = [e for e in in_category if (e.get("prompt") or "") == prompt]
        else:
            filtered = in_category

        if not filtered:
            empty_stats = {
                "total_evaluations": 0,
                "by_category_breakdown": by_category_breakdown,
                "filtered_category": category,
                "prompts_for_category": prompts_for_category,
            }
            if prompt:
                empty_stats["filtered_prompt"] = prompt
                empty_stats["evaluations"] = []
            return {"stats": empty_stats}
        stats = _compute_marketing_stats(filtered)
    else:
        stats = _compute_marketing_stats(evaluations)

    stats["by_category_breakdown"] = by_category_breakdown
    if category:
        stats["filtered_category"] = category
        stats["prompts_for_category"] = prompts_for_category
    if prompt:
        stats["filtered_prompt"] = prompt
        # Attach the raw matching evals so the frontend can render the image
        # gallery + raw counts view without a second round-trip.
        stats["evaluations"] = filtered

    return {"stats": stats}


def _compute_political_stats(evaluations: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Build the political stats payload for a given list of evaluations."""
    total = len(evaluations)
    stats = {
        "total_evaluations": total,
        "by_media_type": {},
        "by_tier": {},
        "by_category": {},
        "stance_distribution": {},
        "sentiment_distribution": {},
        "framing_distribution": {},
        "argument_balance_distribution": {},
    }

    for eval in evaluations:
        mt = eval.get("media_type", "unknown")
        stats["by_media_type"][mt] = stats["by_media_type"].get(mt, 0) + 1

        tier = eval.get("tier", "unknown")
        stats["by_tier"][tier] = stats["by_tier"].get(tier, 0) + 1

        cat = eval.get("category") or "N/A"
        stats["by_category"][cat] = stats["by_category"].get(cat, 0) + 1

        sc = eval.get("stance_code")
        if sc is not None:
            label = BIAS_CODES["stance"].get(sc, "Unknown")
            stats["stance_distribution"][label] = stats["stance_distribution"].get(label, 0) + 1

        sentc = eval.get("sentiment_code")
        if sentc is not None:
            label = BIAS_CODES["sentiment"].get(sentc, "Unknown")
            stats["sentiment_distribution"][label] = stats["sentiment_distribution"].get(label, 0) + 1

        fc = eval.get("framing_code")
        if fc is not None:
            label = BIAS_CODES["framing"].get(fc, "Unknown")
            stats["framing_distribution"][label] = stats["framing_distribution"].get(label, 0) + 1

        ab = eval.get("argument_balance_code")
        if ab is not None:
            label = BIAS_CODES["argument_balance"].get(ab, "Unknown")
            stats["argument_balance_distribution"][label] = stats["argument_balance_distribution"].get(label, 0) + 1

    return stats


@app.get("/api/political-stats")
async def get_political_stats(category: Optional[str] = None, prompt: Optional[str] = None):
    """Get statistical analysis of political bias evaluations.

    Mirrors ``/api/stats``: optional ``category`` (topic) filter, optional
    ``prompt`` filter that composes with it, plus a ``by_category_breakdown``
    always included and a ``prompts_for_category`` list when a topic is set.
    """
    evaluations = []

    if supabase:
        try:
            result = supabase.table("political_evaluations").select("*").execute()
            evaluations = result.data
        except:
            evaluations = in_memory_storage.get("political_evaluations", [])
    else:
        evaluations = in_memory_storage.get("political_evaluations", [])

    if not evaluations:
        return {"message": "No evaluations yet", "stats": {}}

    by_category_breakdown: Dict[str, Any] = {}
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for ev in evaluations:
        key = ev.get("category") or "N/A"
        grouped.setdefault(key, []).append(ev)
    for cat, items in grouped.items():
        by_category_breakdown[cat] = _compute_political_stats(items)

    prompts_for_category: Optional[List[Dict[str, Any]]] = None
    if category:
        in_category = [e for e in evaluations if (e.get("category") or "N/A") == category]
        prompt_counts: Dict[str, int] = {}
        for e in in_category:
            p = e.get("prompt") or ""
            prompt_counts[p] = prompt_counts.get(p, 0) + 1
        prompts_for_category = [
            {"prompt": p, "count": c}
            for p, c in sorted(prompt_counts.items(), key=lambda kv: (-kv[1], kv[0]))
        ]

        if prompt:
            filtered = [e for e in in_category if (e.get("prompt") or "") == prompt]
        else:
            filtered = in_category

        if not filtered:
            empty_stats = {
                "total_evaluations": 0,
                "by_category_breakdown": by_category_breakdown,
                "filtered_category": category,
                "prompts_for_category": prompts_for_category,
            }
            if prompt:
                empty_stats["filtered_prompt"] = prompt
                empty_stats["evaluations"] = []
            return {"stats": empty_stats}
        stats = _compute_political_stats(filtered)
    else:
        stats = _compute_political_stats(evaluations)

    stats["by_category_breakdown"] = by_category_breakdown
    if category:
        stats["filtered_category"] = category
        stats["prompts_for_category"] = prompts_for_category
    if prompt:
        stats["filtered_prompt"] = prompt
        stats["evaluations"] = filtered

    return {"stats": stats}


def _label_or_blank(scheme: str, code) -> str:
    if code is None:
        return ""
    return BIAS_CODES.get(scheme, {}).get(code, "")


def _build_marketing_workbook(evaluations: List[Dict[str, Any]]):
    """Return an openpyxl Workbook of marketing evaluations (raw + decoded labels).

    Rows where the annotator marked ``has_human`` as False have the
    representation/attribute fields filled with ``N/A`` instead of left
    blank — the bias schemes don't apply when there's no person, so this
    distinguishes "intentionally not applicable" from "missing data".
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Marketing Evaluations"

    headers = [
        "evaluation_id", "content_id", "user_id", "evaluated_at",
        "media_type", "area_type", "category", "model_name", "tier",
        "prompt", "content_url",
        "has_human", "human_count",
        "gender_code", "gender_label",
        "race_code", "race_label",
        "age_code", "age_label",
        "occupation_code", "occupation_label",
        "diversity_code", "diversity_label",
        "activity_code", "activity_label",
        "setting_code", "setting_label",
        "appearance_emphasis_code", "appearance_emphasis_label",
        "performance_emphasis_code", "performance_emphasis_label",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    NA = "N/A"

    def code(value, no_human: bool):
        if no_human:
            return NA
        return value

    def label(scheme: str, value, no_human: bool):
        if no_human:
            return NA
        return _label_or_blank(scheme, value)

    for ev in evaluations:
        no_human = ev.get("has_human") is False

        ws.append([
            ev.get("id"), ev.get("content_id"), ev.get("user_id"), ev.get("evaluated_at"),
            ev.get("media_type"), ev.get("area_type"), ev.get("category"),
            ev.get("model_name"), ev.get("tier"),
            ev.get("prompt"), ev.get("content_url"),
            ev.get("has_human"),
            code(ev.get("human_count"), no_human),
            code(ev.get("gender_code"), no_human), label("gender", ev.get("gender_code"), no_human),
            code(ev.get("race_code"), no_human), label("race", ev.get("race_code"), no_human),
            code(ev.get("age_code"), no_human), label("age", ev.get("age_code"), no_human),
            code(ev.get("occupation_code"), no_human), label("occupation", ev.get("occupation_code"), no_human),
            code(ev.get("diversity_code"), no_human), label("diversity", ev.get("diversity_code"), no_human),
            code(ev.get("activity_code"), no_human), label("activity", ev.get("activity_code"), no_human),
            code(ev.get("setting_code"), no_human), label("setting", ev.get("setting_code"), no_human),
            code(ev.get("appearance_emphasis_code"), no_human),
            label("appearance_emphasis", ev.get("appearance_emphasis_code"), no_human),
            code(ev.get("performance_emphasis_code"), no_human),
            label("performance_emphasis", ev.get("performance_emphasis_code"), no_human),
        ])

    return wb


def _build_political_workbook(evaluations: List[Dict[str, Any]]):
    """Return an openpyxl Workbook of political evaluations (raw + decoded labels)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Political Evaluations"

    headers = [
        "evaluation_id", "content_id", "user_id", "evaluated_at",
        "media_type", "area_type", "category", "model_name", "tier",
        "prompt", "content_url",
        "stance_code", "stance_label",
        "sentiment_code", "sentiment_label",
        "framing_code", "framing_label",
        "argument_balance_code", "argument_balance_label",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for ev in evaluations:
        ws.append([
            ev.get("id"), ev.get("content_id"), ev.get("user_id"), ev.get("evaluated_at"),
            ev.get("media_type"), ev.get("area_type"), ev.get("category"),
            ev.get("model_name"), ev.get("tier"),
            ev.get("prompt"), ev.get("content_url"),
            ev.get("stance_code"), _label_or_blank("stance", ev.get("stance_code")),
            ev.get("sentiment_code"), _label_or_blank("sentiment", ev.get("sentiment_code")),
            ev.get("framing_code"), _label_or_blank("framing", ev.get("framing_code")),
            ev.get("argument_balance_code"), _label_or_blank("argument_balance", ev.get("argument_balance_code")),
        ])

    return wb


@app.get("/api/export/{area_type}")
async def export_evaluations(
    area_type: str,
    category: Optional[str] = None,
    prompt: Optional[str] = None,
):
    """Download raw evaluation data as an .xlsx workbook.

    ``area_type`` must be ``marketing``, ``political``, or ``all``. Optional
    ``category`` and ``prompt`` filters compose to narrow the rows that are
    written to the workbook.
    """
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    area_type = (area_type or "").lower()
    if area_type not in ("marketing", "political", "all"):
        raise HTTPException(status_code=400, detail="Invalid area_type. Use 'marketing', 'political', or 'all'.")

    # Pull marketing evaluations
    marketing_evals: List[Dict[str, Any]] = []
    if area_type in ("marketing", "all"):
        if supabase:
            try:
                result = supabase.table("evaluations").select("*").order("evaluated_at", desc=True).execute()
                marketing_evals = result.data or []
            except Exception:
                marketing_evals = list(in_memory_storage.get("evaluations", []))
        else:
            marketing_evals = list(in_memory_storage.get("evaluations", []))
        if category:
            marketing_evals = [e for e in marketing_evals if (e.get("category") or "N/A") == category]
        if prompt:
            marketing_evals = [e for e in marketing_evals if (e.get("prompt") or "") == prompt]

    # Pull political evaluations
    political_evals: List[Dict[str, Any]] = []
    if area_type in ("political", "all"):
        if supabase:
            try:
                result = supabase.table("political_evaluations").select("*").order("evaluated_at", desc=True).execute()
                political_evals = result.data or []
            except Exception:
                political_evals = list(in_memory_storage.get("political_evaluations", []))
        else:
            political_evals = list(in_memory_storage.get("political_evaluations", []))
        if category:
            political_evals = [e for e in political_evals if (e.get("category") or "N/A") == category]
        if prompt:
            political_evals = [e for e in political_evals if (e.get("prompt") or "") == prompt]

    # Build the workbook(s)
    if area_type == "marketing":
        wb = _build_marketing_workbook(marketing_evals)
        filename_stem = "marketing_evaluations"
    elif area_type == "political":
        wb = _build_political_workbook(political_evals)
        filename_stem = "political_evaluations"
    else:
        # area_type == "all" — combine both into one workbook with two sheets
        from openpyxl import Workbook
        wb = Workbook()
        # Replace the default sheet with marketing sheet
        default_ws = wb.active
        wb.remove(default_ws)
        wb_m = _build_marketing_workbook(marketing_evals)
        wb_p = _build_political_workbook(political_evals)
        # Copy sheets across
        for src_wb in (wb_m, wb_p):
            src_ws = src_wb.active
            new_ws = wb.create_sheet(title=src_ws.title)
            for row in src_ws.iter_rows(values_only=True):
                new_ws.append(row)
        filename_stem = "all_evaluations"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    suffix = f"_{category}" if category else ""
    filename = f"{filename_stem}{suffix}_{timestamp}.xlsx"
    # Sanitize filename — replace anything that's not alnum/dot/dash/underscore
    safe = "".join(c if c.isalnum() or c in "._-" else "_" for c in filename)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe}"'},
    )


@app.get("/api/image-proxy")
async def image_proxy(url: str):
    """Proxy endpoint to serve images and handle CORS/SSL issues with retry logic"""
    max_retries = 3
    retry_delay = 0.5
    last_error = None
    
    for attempt in range(max_retries):
        try:
            # Try with SSL verification first
            verify_ssl = attempt < 2  # Disable SSL verification on last attempt
            async with httpx.AsyncClient(verify=verify_ssl) as client:
                response = await client.get(
                    url, 
                    follow_redirects=True, 
                    timeout=httpx.Timeout(20.0, connect=10.0),
                    headers={
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                        'Accept': 'image/*',
                        'Accept-Language': 'en-US,en;q=0.9',
                        'Cache-Control': 'no-cache'
                    }
                )
                response.raise_for_status()
                
                # Get content type, default to image/jpeg if not provided
                content_type = response.headers.get('content-type', 'image/jpeg')
                
                # Ensure we have image content
                if 'image' not in content_type.lower():
                    print(f"Warning: Received non-image content type: {content_type}")
                    if attempt < max_retries - 1:
                        await asyncio.sleep(retry_delay)
                        continue
                
                # Check if we got valid content
                if not response.content or len(response.content) < 100:
                    print(f"Warning: Image content too small: {len(response.content) if response.content else 0} bytes")
                    if attempt < max_retries - 1:
                        await asyncio.sleep(retry_delay)
                        continue
                
                return {
                    "success": True,
                    "data": base64.b64encode(response.content).decode('utf-8'),
                    "content_type": content_type
                }
        except httpx.TimeoutException as e:
            last_error = f"Timeout: {str(e)}"
            print(f"Image proxy timeout (attempt {attempt + 1}/{max_retries}): {e}")
        except httpx.HTTPStatusError as e:
            last_error = f"HTTP {e.response.status_code}: {str(e)}"
            print(f"Image proxy HTTP error (attempt {attempt + 1}/{max_retries}): {e}")
        except Exception as e:
            last_error = str(e)
            print(f"Image proxy error (attempt {attempt + 1}/{max_retries}): {e}")
        
        # Wait before retrying
        if attempt < max_retries - 1:
            await asyncio.sleep(retry_delay * (attempt + 1))
    
    # If all retries failed, raise error
    print(f"Image proxy failed after {max_retries} attempts. Last error: {last_error}")
    raise HTTPException(status_code=502, detail=f"Failed to fetch image. Last error: {last_error}")


@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "supabase_connected": supabase is not None,
        "cloudinary_configured": os.getenv("CLOUDINARY_CLOUD_NAME") not in [None, "your_cloud_name"]
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
